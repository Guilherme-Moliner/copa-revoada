#!/usr/bin/env python3
"""
Copa Revoada — build do site.

Lê dados/COPA_REVOADA_planilha.xlsx, junta com as fotos de fotos/ e o logo de
assets/, e escreve index.html a partir de src/app.template.html.

    python3 scripts/build.py

Nada aqui depende de internet. As únicas dependências são openpyxl e Pillow.
"""

import base64
import io
import json
import os
import sys
import unicodedata
import re

try:
    import openpyxl
    from PIL import Image, ImageEnhance, ImageOps
except ImportError:
    sys.exit("Faltam dependências. Rode:  pip install openpyxl pillow")

# opencv é opcional: com ele o recorte do retrato acha o rosto; sem ele, cai no
# enquadramento por proporção. O build funciona dos dois jeitos.
try:
    import cv2
    import numpy as np
    TEM_ROSTO = os.path.isfile(
        os.path.join(cv2.data.haarcascades, "haarcascade_frontalface_default.xml"))
except ImportError:
    TEM_ROSTO = False

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PLANILHA = os.path.join(RAIZ, "dados", "COPA_REVOADA_planilha.xlsx")
LEGADO = os.path.join(RAIZ, "dados", "ranking-legado.json")
TEMPLATE = os.path.join(RAIZ, "src", "app.template.html")
LOGO = os.path.join(RAIZ, "assets", "logo.png")
FOTOS = os.path.join(RAIZ, "fotos")
ASSETS = os.path.join(RAIZ, "assets")
IMG = os.path.join(RAIZ, "img")
SAIDA = os.path.join(RAIZ, "index.html")

PLAYLIST = "PLSnrz0oA5cB49hSu2rMMCCtbXZOK6l9hq"

# Fonte de verdade dos dados. Com PLANILHA_URL definida, o build lê da planilha
# online em vez do arquivo do repositório — que passa a ser só uma cópia de
# segurança. O repositório guarda código; a planilha guarda dado operacional.
#
# Aceita duas formas:
#   1. o Apps Script:  .../exec?acao=planilha   → devolve o .xlsx em base64
#   2. o export direto: .../export?format=xlsx  → só funciona com planilha pública
#
# A do Apps Script é a boa: a planilha continua privada e quem busca é o script,
# rodando com a credencial do dono.
PLANILHA_URL = os.environ.get("PLANILHA_URL", "").strip()

# Endereço do Apps Script publicado como Web App. É por ele que o app grava e
# lê os lances marcados. Sem ele, a Análise do jogo funciona só na memória da
# aba. Ver docs/planilha-colaborativa.md.
LANCES_URL = os.environ.get("LANCES_URL", "").strip()

# Nos três primeiros jogos ninguém anotou gol nem assistência por jogador, só
# quem foi campeão. Decisão do grupo: esses jogos contam título e nada mais.
# É o que faz cada um ter 10 jogos válidos e 13 presenças.
# Esvaziar este conjunto devolve a contagem antiga de 13.
SO_TITULO = {"2020-12", "2021-07", "2021-12"}

EXTS = (".jpg", ".jpeg", ".png", ".webp")
AVISOS = []

# Arquivos de assets/ cujo nome não bate com o id da planilha. Só entram aqui
# os casos em que a correspondência é evidente (erro de digitação ou apelido do
# time). Nome duvidoso não é chutado: vira aviso no fim do build.
APELIDO_JOGADOR = {
    "danilin": "danilim",
    "leobittencour": "leobittencourt",
    "be": "correa",          # Be / Be Correa / Bernardo C
    "bemarucco": "marucco",  # Be Marucco / Bernardo M
    "dede": "andre",         # Dedé é o André
    "vitim": "kretzer",      # Vitim e Kretzer são a mesma pessoa
}
ESCUDO_TIME = {
    "borussetalogo": "borussia22",   # agora é o BVB mesmo
    "bocalogo": "boca23",            # CABJ
    "mancitylogo": "city23",         # Mano Cityfodo
    "pegueisuagatalogo": "psg20",    # Peguei Sua Gata
    "criseumafclogo": "preto24",
    "dboafclogo": "branco24",
    "dentrofclogo": "dentro26",
    "ferroviagralogo": "ferroviagra26",
    "jumentuslogo": "jumentus25",
    "liverpoollogo": "liverpool22",
    "milanblogo": "milanbe25",
}
FOTO_TIME = {
    "borusseta": "borussia22",
    "citifodo": "city23",
    "criseuma": "preto24",       # Criseúma FC é o time preto de 2024
    "dboafc": "branco24",        # D Boa FC é o time branco de 2024
    "dentofc2026": "dentro26",
    "ferroviagra2026": "ferroviagra26",
    "jumentos": "jumentus25",
    "liverpool": "liverpool22",
    "milanb": "milanbe25",
}
# Lançamento errado na planilha antiga: fica lá para conferência, mas não entra
# no site nem conta estatística. Tirar daqui devolve a pessoa ao ar.
EXCLUIR_DO_SITE = {
    "derek", "davi", "dereck", "matheus", "luigi", "dereka", "joao", "miniderek",
}
# Não são foto de jogador nem de time; ficam de fora do site até alguém dizer
# o que são. Listados aqui só para não aparecerem como "arquivo sem uso".
IGNORAR_ASSET = {"logo", "logo-azul", "logo-branco", "semperfil", "dede2",
                 "pegueisuagatalogo2"}

# Retrato de reserva para quem ainda não tem foto. Melhor uma silhueta do que
# duas iniciais soltas — e mantém a mesma moldura em todas as telas.
SEM_PERFIL = "semperfil"


def aviso(msg):
    AVISOS.append(msg)


def slug(s):
    s = unicodedata.normalize("NFD", str(s)).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]", "", s)


def linhas(ws, cabecalho_em):
    """Devolve dicts a partir da linha de cabeçalho informada (1-based)."""
    hdr = [str(c.value).strip() if c.value else "" for c in ws[cabecalho_em]]
    for row in ws.iter_rows(min_row=cabecalho_em + 1, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        yield {hdr[i]: row[i] for i in range(len(hdr)) if hdr[i]}


def acha_cabecalho(ws, primeira_coluna):
    """A planilha tem uma nota na linha 1; o cabeçalho pode estar na 1 ou na 3."""
    for r in (1, 2, 3, 4):
        v = ws.cell(r, 1).value
        if v and str(v).strip() == primeira_coluna:
            return r
    raise SystemExit(f"Não achei o cabeçalho '{primeira_coluna}' na aba {ws.title}")


def _id_video(v):
    """Aceita a URL inteira ou só o id do YouTube."""
    t = str(v or "").strip()
    if not t:
        return ""
    m = re.search(r"(?:v=|youtu\.be/|embed/|shorts/)([\w-]{6,})", t)
    return m.group(1) if m else t


def num(v, padrao=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return padrao


def sim(v):
    return str(v).strip().upper() in ("SIM", "S", "1", "TRUE", "X")


USADOS = set()


def acha_origem(base, pastas=(FOTOS, ASSETS)):
    """Procura <base>.<ext> nas pastas dadas, na ordem."""
    for pasta in pastas:
        for ext in EXTS:
            caminho = os.path.join(pasta, base + ext)
            if os.path.exists(caminho):
                USADOS.add(base.lower())
                return caminho
    return ""


def _destino(sub):
    caminho = os.path.join(IMG, sub)
    os.makedirs(caminho, exist_ok=True)
    return caminho


ROSTOS_ACHADOS = []
ROSTOS_PERDIDOS = []


def acha_rosto(im):
    """(centro_x, centro_y, largura, altura) do maior rosto, ou None."""
    if not TEM_ROSTO:
        return None
    cinza = np.array(im.convert("L"))
    # equalizar ajuda bastante em foto de jogo, com sol estourado ou sombra
    cinza = cv2.equalizeHist(cinza)
    alt, larg = cinza.shape
    lado_menor = min(alt, larg)

    def plausivel(f):
        """Rosto de retrato fica na metade de cima e tem tamanho de rosto.
        Sem isso, a passada mais frouxa acha 'rosto' em grama e alambrado."""
        x, y, w, h = f
        cy = y + h / 2
        return (cy < alt * .72                      # não fica lá embaixo
                and .05 < w / larg < .75            # nem minúsculo nem a foto toda
                and .5 < w / h < 1.6)               # rosto é quase quadrado

    # afrouxa aos poucos; a foto de corpo inteiro só cai nas últimas
    for minimo, vizinhos in ((max(40, lado_menor // 8), 6),
                             (max(28, lado_menor // 16), 5),
                             (max(20, lado_menor // 22), 4)):
        for arquivo in ("haarcascade_frontalface_default.xml",
                        "haarcascade_frontalface_alt2.xml",
                        "haarcascade_profileface.xml"):
            cascata = cv2.CascadeClassifier(cv2.data.haarcascades + arquivo)
            achados = [f for f in cascata.detectMultiScale(
                cinza, scaleFactor=1.06, minNeighbors=vizinhos,
                minSize=(minimo, minimo))if plausivel(f)]
            if achados:
                x, y, w, h = max(achados, key=lambda f: f[2] * f[3])
                return x + w / 2, y + h / 2, w, h
    return None


def deriva_retrato(origem, nome):
    """Recorta quadrado em cima do rosto e grava JPEG leve e tratado."""
    saida = os.path.join(_destino("jogadores"), nome + ".jpg")
    im = Image.open(origem).convert("RGB")

    achado = acha_rosto(im)
    if achado:
        cx, cy, _, fh = achado
        # o rosto fica ocupando ~46% da altura do quadro, com ar em cima
        lado = int(min(min(im.size), fh * 2.2))
        x, y = int(cx - lado / 2), int(cy - lado * .42)
        ROSTOS_ACHADOS.append(nome)
    else:
        lado = min(im.size)
        x, y = (im.width - lado) // 2, int((im.height - lado) * .18)
        ROSTOS_PERDIDOS.append(nome)

    x = max(0, min(x, im.width - lado))
    y = max(0, min(y, im.height - lado))
    im = im.crop((x, y, x + lado, y + lado))
    im = im.resize((420, 420), Image.LANCZOS)

    # tratamento: nivela exposição, dá contraste e um pouco de nitidez
    im = ImageOps.autocontrast(im, cutoff=(1, 2))
    im = ImageEnhance.Color(im).enhance(1.08)
    im = ImageEnhance.Sharpness(im).enhance(1.4)

    im.save(saida, "JPEG", quality=86, optimize=True, progressive=True)
    return "img/jogadores/%s.jpg" % nome


FUNDO_TIRADO = []
VAZOU = []


def tira_fundo(im):
    """Escudo que veio em fundo chapado sai com fundo transparente.

    Só age quando (a) a imagem não tem transparência de verdade e (b) a borda
    é de uma cor só. O preenchimento parte das quatro quinas, então branco de
    dentro do escudo não é apagado junto.
    """
    im = im.convert("RGBA")
    if not TEM_ROSTO:                      # sem opencv não tem floodFill
        return im
    a = np.array(im)
    if (a[:, :, 3] < 250).mean() > .06:    # já tem alpha; não mexe
        return im
    rgb = a[:, :, :3]
    alt, larg = a.shape[:2]
    # A condição é que as QUATRO QUINAS combinem entre si — são elas que dão a
    # semente. Exigir a borda inteira de uma cor só era severo demais: no BVB o
    # círculo amarelo encosta na margem, e o escudo ficava com o quadrado branco.
    quinas = np.array([rgb[0, 0], rgb[0, -1], rgb[-1, 0], rgb[-1, -1]], np.int16)
    if (quinas.max(axis=0) - quinas.min(axis=0)).max() > 26:
        return im
    mascara = np.zeros((alt + 2, larg + 2), np.uint8)
    bgr = np.ascontiguousarray(rgb[:, :, ::-1])
    tol = (20, 20, 20)
    # FIXED_RANGE compara cada pixel com a COR DA SEMENTE. Sem essa flag o
    # openCV compara com o pixel vizinho, e aí um fundo com leve degradê deixa
    # o preenchimento caminhar para dentro do desenho: era o que estava
    # comendo o escudo do Peguei Sua Gata (sobravam 7% de pixel opaco).
    modo = 4 | (255 << 8) | cv2.FLOODFILL_MASK_ONLY | cv2.FLOODFILL_FIXED_RANGE
    for semente in ((0, 0), (larg - 1, 0), (0, alt - 1), (larg - 1, alt - 1)):
        cv2.floodFill(bgr, mascara, semente, 0, tol, tol, modo)
    fundo = mascara[1:-1, 1:-1].astype(bool)
    if fundo.mean() < .04:                 # não achou fundo que valha a pena
        return im
    # trava: logo claro sobre fundo claro faz o preenchimento vazar para dentro
    # do desenho. Se sobrou pouca coisa opaca, é porque comeu o escudo junto.
    if 1 - fundo.mean() < .12:
        VAZOU.append(True)
        return im
    a[fundo, 3] = 0
    return Image.fromarray(a, "RGBA")


def deriva_escudo(origem, nome):
    """Escudo mantém transparência; PNG pequeno."""
    saida = os.path.join(_destino("times"), nome + "-escudo.png")
    bruto = Image.open(origem).convert("RGBA")
    im = tira_fundo(bruto)
    if im is not bruto:
        vazio_antes = (np.array(bruto)[:, :, 3] < 250).mean()
        vazio_depois = (np.array(im)[:, :, 3] < 250).mean()
        if vazio_depois - vazio_antes > .02:
            FUNDO_TIRADO.append(nome)
    caixa = im.getbbox()
    if caixa:
        im = im.crop(caixa)
    im.thumbnail((320, 320), Image.LANCZOS)
    im.save(saida, "PNG", optimize=True)
    return "img/times/%s-escudo.png" % nome


def deriva_foto_time(origem, nome):
    saida = os.path.join(_destino("times"), nome + "-foto.jpg")
    im = Image.open(origem).convert("RGB")
    im.thumbnail((1100, 1100), Image.LANCZOS)
    im.save(saida, "JPEG", quality=80, optimize=True, progressive=True)
    return "img/times/%s-foto.jpg" % nome


SEM_FOTO = []


def foto_de(pid):
    origem = acha_origem(pid)
    if not origem:
        for arquivo, alvo in APELIDO_JOGADOR.items():
            if alvo == pid:
                origem = acha_origem(arquivo)
                break
    if origem:
        return deriva_retrato(origem, pid)
    SEM_FOTO.append(pid)
    return ""


def retrato_reserva():
    """A silhueta que entra no lugar de quem ainda não tem foto."""
    origem = acha_origem(SEM_PERFIL, (ASSETS,))
    if not origem:
        return ""
    saida = os.path.join(_destino("jogadores"), "_sem-perfil.jpg")
    im = Image.open(origem).convert("RGB")
    lado = min(im.size)
    x, y = (im.width - lado) // 2, (im.height - lado) // 2
    im = im.crop((x, y, x + lado, y + lado)).resize((420, 420), Image.LANCZOS)
    im.save(saida, "JPEG", quality=84, optimize=True, progressive=True)
    return "img/jogadores/_sem-perfil.jpg"


def imagens_do_time(tid):
    """Devolve (escudo, foto) já otimizados, ou strings vazias."""
    escudo = foto = ""
    if tid == "psg21":                       # os dois anos usam o mesmo escudo
        tid_escudo = "psg20"
    else:
        tid_escudo = tid
    for arquivo, alvo in ESCUDO_TIME.items():
        if alvo == tid_escudo:
            o = acha_origem(arquivo, (ASSETS,))
            if o:
                escudo = deriva_escudo(o, tid)
            break
    for arquivo, alvo in FOTO_TIME.items():
        if alvo == tid:
            o = acha_origem(arquivo, (ASSETS,))
            if o:
                foto = deriva_foto_time(o, tid)
            break
    return escudo, foto


def assets_sem_uso():
    """Arquivo de imagem em assets/ que o build não soube onde encaixar."""
    sobra = []
    for f in sorted(os.listdir(ASSETS)):
        base, ext = os.path.splitext(f)
        if ext.lower() not in EXTS or base.lower() in IGNORAR_ASSET:
            continue
        if base.lower() not in USADOS:
            sobra.append(f)
    return sobra


def logo_b64(px=300):
    im = Image.open(LOGO)
    im = im.crop(im.getbbox())
    im.thumbnail((px, px), Image.LANCZOS)
    alpha = im.split()[3] if im.mode == "RGBA" else None
    rgb = im.convert("RGB").quantize(colors=4).convert("RGB")
    if alpha:
        rgb.putalpha(alpha)
    buf = io.BytesIO()
    rgb.save(buf, "PNG", optimize=True)
    return base64.b64encode(buf.getvalue()).decode()


class _Celula:
    __slots__ = ("value",)

    def __init__(self, valor):
        self.value = valor


class _Aba:
    """Finge ser uma worksheet do openpyxl, só com o que o build usa.

    O Apps Script devolve as abas como listas de listas. Envolver isso numa
    casca com a mesma cara do openpyxl evita reescrever todos os leitores —
    eles continuam fazendo ws.cell(), ws[3] e ws.iter_rows() sem saber de nada.
    """

    def __init__(self, titulo, linhas):
        self.title = titulo
        self._linhas = linhas
        self.max_row = len(linhas)
        self.max_column = max((len(l) for l in linhas), default=0)

    def cell(self, linha, coluna, valor=None):
        if 1 <= linha <= self.max_row:
            l = self._linhas[linha - 1]
            if 1 <= coluna <= len(l):
                return _Celula(l[coluna - 1])
        return _Celula(None)

    def __getitem__(self, linha):
        return [self.cell(linha, c + 1) for c in range(self.max_column)]

    def iter_rows(self, min_row=1, max_row=None, values_only=False):
        fim = max_row or self.max_row
        for i in range(min_row, fim + 1):
            l = list(self._linhas[i - 1]) if i <= self.max_row else []
            l += [None] * (self.max_column - len(l))
            yield tuple(l) if values_only else [_Celula(v) for v in l]


class _Planilha:
    def __init__(self, abas):
        self._abas = {k: _Aba(k, v) for k, v in abas.items()}
        self.sheetnames = list(abas.keys())

    def __getitem__(self, nome):
        return self._abas[nome]


def baixa_planilha():
    """Baixa a planilha publicada e devolve o caminho de um arquivo local.

    Guarda uma cópia em dados/ para o build continuar funcionando offline se
    o Sheets estiver fora do ar na hora do deploy.
    """
    import urllib.request
    copia = os.path.join(RAIZ, "dados", "COPA_REVOADA_baixada.xlsx")
    copia_json = os.path.join(RAIZ, "dados", "COPA_REVOADA_baixada.json")
    try:
        print(f"  baixando a planilha de {PLANILHA_URL[:60]}...")
        req = urllib.request.Request(PLANILHA_URL, headers={"User-Agent": "copa-revoada"})
        with urllib.request.urlopen(req, timeout=60) as r:
            dados = r.read()
        if dados[:1] == b"{":
            pacote = json.loads(dados.decode("utf-8"))
            if not pacote.get("ok"):
                raise ValueError(pacote.get("erro", "o script recusou"))
            abas = pacote["abas"]
            with open(copia_json, "w", encoding="utf-8") as f:
                json.dump(abas, f, ensure_ascii=False)
            print(f"  planilha lida do Google Sheets — {len(abas)} abas, "
                  f"{sum(len(v) for v in abas.values())} linhas")
            return _Planilha(abas)
        if len(dados) < 5000 or dados[:2] != b"PK":
            raise ValueError("a resposta não parece planilha — confira a URL")
        with open(copia, "wb") as f:
            f.write(dados)
        print(f"  planilha baixada — {len(dados)//1024} KB")
        return copia
    except Exception as e:
        if os.path.exists(copia_json):
            aviso(f"não deu para ler a planilha online ({e}); usando a última cópia baixada")
            return _Planilha(json.load(open(copia_json, encoding="utf-8")))
        if os.path.exists(copia):
            aviso(f"não deu para ler a planilha online ({e}); usando a última cópia baixada")
            return copia
        aviso(f"não deu para ler a planilha online ({e}); usando o arquivo do repositório")
        return PLANILHA


def main():
    origem = baixa_planilha() if PLANILHA_URL else PLANILHA
    wb = origem if isinstance(origem, _Planilha) else openpyxl.load_workbook(origem, data_only=True)

    # ---------------- jogadores ----------------
    ws = wb["JOGADORES"]
    jogadores = []
    for r in linhas(ws, acha_cabecalho(ws, "id")):
        pid = str(r.get("id") or "").strip()
        if not pid or pid in EXCLUIR_DO_SITE:
            continue
        jogadores.append({
            "id": pid,
            "apelido": str(r.get("apelido") or pid).strip(),
            # coluna nova da planilha: o apelido de zoeira que aparece no perfil.
            # aceita alguns nomes de cabeçalho porque quem preenche escolhe o seu
            "apelido2": next((str(r[k]).strip() for k in
                              ("apelido_engracado", "apelido2", "apelido_zoeira",
                               "apelido_engraçado", "novo_apelido", "apelido_novo")
                              if r.get(k)), ""),
            "nome": str(r.get("nome_completo") or "").strip(),
            "pos": str(r.get("posicao") or "").strip().upper(),
            "numero": num(r.get("numero_camisa"), 0),
            "foto": foto_de(pid),
            "jogos": 0, "gols": 0, "assist": 0, "titulos": 0, "contra": 0,
            "presencas": 0,
        })
    reserva = retrato_reserva()
    if reserva:
        for p in jogadores:
            if not p["foto"]:
                p["foto"] = reserva
                p["semFoto"] = True
    porId = {p["id"]: p for p in jogadores}

    # ---------------- times ----------------
    ws = wb["TIMES"]
    times = []
    for r in linhas(ws, acha_cabecalho(ws, "id")):
        tid = str(r.get("id") or "").strip()
        if not tid:
            continue
        escudo, foto = imagens_do_time(tid)
        # a planilha pode apontar um arquivo à mão; quando aponta, ela manda
        manual = str(r.get("escudo_arquivo") or "").strip()
        if manual:
            o = acha_origem(os.path.splitext(manual)[0], (ASSETS,))
            if o:
                escudo = deriva_escudo(o, tid)
            else:
                aviso(f"TIMES {tid}: escudo_arquivo '{manual}' não existe em assets/")
        times.append({
            "id": tid,
            "nome": str(r.get("nome") or tid).strip(),
            "temporada": num(r.get("temporada")),
            "sigla": str(r.get("sigla") or tid[:3]).strip().upper(),
            "c1": str(r.get("cor_1") or "#12A150").strip(),
            "c2": str(r.get("cor_2") or "#F2B01E").strip(),
            "escudo": escudo,
            "foto": foto,
        })
    times.sort(key=lambda t: (t["temporada"], t["id"]))
    timeIds = {t["id"] for t in times}

    # ---------------- jogos ----------------
    ws = wb["JOGOS"]
    jogos = []
    for r in linhas(ws, acha_cabecalho(ws, "id")):
        gid = str(r.get("id") or "").strip()
        if not gid:
            continue
        jogos.append({
            "id": gid,
            "temporada": num(r.get("temporada")),
            "mes": str(r.get("mes") or "").strip(),
            "dia": str(r.get("dia") or "").strip(),
            "casa": str(r.get("time_casa") or "").strip(),
            "gc": num(r.get("gols_casa")),
            "fora": str(r.get("time_fora") or "").strip(),
            "gf": num(r.get("gols_fora")),
            "estadio": str(r.get("estadio") or "").strip(),
            "campeao": str(r.get("campeao_time_id") or "").strip(),
            "video": _id_video(next((r[k] for k in
                     ("video_youtube", "link_yyt", "link_yt", "link_youtube",
                      "video", "youtube", "link") if r.get(k)), "")),
        })
    jogos.sort(key=lambda g: g["id"])
    jogoIds = {g["id"] for g in jogos}

    for g in jogos:
        for campo in ("casa", "fora"):
            if g[campo] and g[campo] not in timeIds:
                aviso(f"JOGOS {g['id']}: time '{g[campo]}' não existe na aba TIMES")

    # ---------------- escalações ----------------
    escalacoes = []
    if "ESCALACOES" in wb.sheetnames:
        ws = wb["ESCALACOES"]
        for r in linhas(ws, acha_cabecalho(ws, "jogo_id")):
            jid = str(r.get("jogo_id") or "").strip()
            pid = str(r.get("jogador_id") or "").strip()
            if not jid or not pid:
                continue
            if jid.upper().startswith("EXEMPLO") or pid.upper().startswith("EXEMPLO"):
                continue
            if jid not in jogoIds:
                aviso(f"ESCALACOES: jogo '{jid}' não existe na aba JOGOS")
                continue
            oculto = pid in EXCLUIR_DO_SITE
            if not oculto and pid not in porId:
                aviso(f"ESCALACOES: jogador '{pid}' não existe na aba JOGADORES")
                continue
            escalacoes.append({
                "jogo": jid, "jogador": pid,
                "time": str(r.get("time_id") or "").strip(),
                "gols": num(r.get("gols")),
                "assist": num(r.get("assistencias")),
                "gols_lancados": r.get("gols") is not None and str(r.get("gols")).strip() != "",
                "campeao": sim(r.get("campeao")),
                "nota": num(r.get("nota_1a5"), 0) or None,
                "numero": num(r.get("numero"), 0) or None,
                # gol contra: conta pro adversário, nunca pro total do jogador
                # o jogador some do site, mas o que ele fez continua contando
                # para o time: gol aconteceu, e esconder o nome não desfaz
                "oculto": oculto,
                "contra": num(r.get("gols_contra")),
                # gol anulado (VAR): vale pro jogador, não vale pro placar
                "no_placar": str(r.get("conta_no_placar") or "SIM").strip().upper() != "NAO",
            })

    # ---------------- desempenho por time ----------------
    # o rostinho da página do time sai daqui: é o julgamento de como a pessoa
    # foi naquele time, não em cada jogo isolado
    desempenho = []
    if "DESEMPENHO" in wb.sheetnames:
        ws = wb["DESEMPENHO"]
        for r in linhas(ws, acha_cabecalho(ws, "time_id")):
            tid = str(r.get("time_id") or "").strip()
            pid = str(r.get("jogador_id") or "").strip()
            nota = num(r.get("nota_1a5"), 0)
            if not tid or not pid or not nota:
                continue
            if tid not in timeIds:
                aviso(f"DESEMPENHO: time '{tid}' não existe na aba TIMES")
                continue
            if pid in EXCLUIR_DO_SITE:
                continue
            if pid not in porId:
                aviso(f"DESEMPENHO: jogador '{pid}' não existe na aba JOGADORES")
                continue
            desempenho.append({
                "time": tid, "jogador": pid,
                "nota": max(1, min(5, nota)),
                "obs": str(r.get("observacao") or "").strip(),
            })

    # ---------------- estatísticas ----------------
    cobertos = {e["jogo"] for e in escalacoes}
    faltando = [g["id"] for g in jogos if g["id"] not in cobertos]
    # os totais só passam a sair das escalações quando TODOS os jogos estiverem
    # lançados; parcial contaria menos jogo do que a pessoa realmente fez
    completo = bool(jogos) and not faltando

    jogos_sem_gol = sorted({e["jogo"] for e in escalacoes if not e["gols_lancados"]})

    # ---------------- placar x gols lançados ----------------
    # é a conferência que a aba CONFERE fazia à mão. Gol contra vai pro outro
    # lado; gol anulado pelo VAR não entra. Sobrou diferença, o build reclama.
    # só pula o jogo em que NINGUÉM teve gol lançado. Jogo com uma ou outra
    # célula vazia continua conferível — 2022-04 e 2022-10 são desses.
    sem_nenhum = {g["id"] for g in jogos
                  if not any(e["gols_lancados"] for e in escalacoes if e["jogo"] == g["id"])}
    for g in jogos:
        if g["id"] in sem_nenhum or not g["casa"] or not g["fora"]:
            continue
        marcou = {g["casa"]: 0, g["fora"]: 0}
        for e in escalacoes:
            if e["jogo"] != g["id"] or e["time"] not in marcou:
                continue
            if e["no_placar"]:
                marcou[e["time"]] += e["gols"]
            if e["contra"]:
                outro = g["fora"] if e["time"] == g["casa"] else g["casa"]
                marcou[outro] += e["contra"]
        if marcou[g["casa"]] != g["gc"] or marcou[g["fora"]] != g["gf"]:
            aviso(f"{g['id']}: as escalações somam {marcou[g['casa']]}x{marcou[g['fora']]} "
                  f"mas o placar da aba JOGOS é {g['gc']}x{g['gf']}")

    if completo:
        for e in escalacoes:
            if e["oculto"]:
                continue          # não entra em estatística individual
            p = porId[e["jogador"]]
            # presença é toda escalação; jogo válido exclui os três primeiros,
            # em que só o título conta
            p["presencas"] += 1
            if e["jogo"] not in SO_TITULO:
                p["jogos"] += 1
                p["gols"] += e["gols"]
                p["assist"] += e["assist"]
                p["contra"] += e["contra"]
            if e["campeao"]:
                p["titulos"] += 1

        validos = [g["id"] for g in jogos if g["id"] not in SO_TITULO]
        print(f"  totais de {len(escalacoes)} escalacoes — {len(validos)} jogos valem "
              f"estatistica, {len(SO_TITULO)} valem so titulo")
        # depois da regra dos três primeiros, o ranking antigo deixa de ser
        # necessário como fonte: os gols saem inteiros das escalações
        ainda_sem = sorted(set(jogos_sem_gol) - SO_TITULO - set(sem_nenhum))
        if ainda_sem:
            aviso("Jogos que valem estatística e ainda têm célula de gol em branco: "
                  + ", ".join(ainda_sem) + ". O placar bate, então provavelmente a "
                  "célula vazia quer dizer zero — mas confirme.")
        leg = {slug(x["id"]): x for x in json.load(open(LEGADO, encoding="utf-8"))}
        difs = []
        for p in jogadores:
            l = leg.get(slug(p["id"]))
            if l and l["jogos"] and abs(l["jogos"] - p["jogos"]) > 2:
                difs.append(f"{p['apelido']} {p['jogos']}x{l['jogos']}")
        if difs:
            aviso("Contagem de jogos ainda diverge do ranking antigo mesmo depois da "
                  "regra dos três primeiros (site x ranking): " + ", ".join(difs) + ".")
    else:
        legado = {slug(x["id"]): x for x in json.load(open(LEGADO, encoding="utf-8"))}
        achou = 0
        for p in jogadores:
            l = legado.get(slug(p["id"]))
            if l:
                p.update(jogos=l["jogos"], gols=l["gols"], presencas=l["jogos"],
                         assist=l["assist"], titulos=l["titulos"])
                achou += 1
        print(f"  totais vindos do ranking antigo ({achou} jogadores) — "
              f"ESCALACOES cobre {len(cobertos)}/{len(jogos)} jogos")
        if faltando:
            aviso("Sem escalação lançada em: " + ", ".join(faltando) +
                  ". Enquanto faltar algum jogo, os totais de cada jogador vêm do ranking "
                  "antigo (as telas de elenco e time já usam as escalações que existem).")

    # ---------------- troféus e premiações ----------------
    # todos são de ouro: é assim que os troféus da copa existem de verdade.
    # o "modelo" diz qual arte o site desenha.
    trofeus = [
        {"id": "campeao", "nome": "Campeão da Temporada", "tier": "ouro", "modelo": "lutador"},
        {"id": "artilheiro", "nome": "Artilheiro", "tier": "ouro", "modelo": "chuteira"},
        {"id": "garcom", "nome": "Garçom", "tier": "ouro", "modelo": "bandeja"},
        {"id": "craque", "nome": "Bola de Ouro", "tier": "ouro", "modelo": "bola"},
        {"id": "paredao", "nome": "Paredão", "tier": "ouro", "modelo": "goleiro"},
        {"id": "piroquinha", "nome": "Piroquinha", "tier": "ouro", "modelo": "piroquinha"},
    ]
    premiacoes = []
    if "TROFEUS" in wb.sheetnames:
        ws = wb["TROFEUS"]
        for r in linhas(ws, acha_cabecalho(ws, "temporada")):
            pid = str(r.get("jogador_id") or "").strip()
            if not pid or str(r.get("jogo_id") or "").upper().startswith("EXEMPLO"):
                continue
            if pid in EXCLUIR_DO_SITE:
                continue
            if pid not in porId:
                aviso(f"TROFEUS: jogador '{pid}' não existe na aba JOGADORES")
                continue
            premiacoes.append({
                "temporada": num(r.get("temporada")),
                "jogo": str(r.get("jogo_id") or "").strip(),
                "trofeu": str(r.get("trofeu_id") or "").strip(),
                "jogador": pid,
                "obs": str(r.get("observacao") or "").strip(),
            })

    # ---------------- vídeos ----------------
    # a aba VIDEOS manda; a coluna video_youtube da aba JOGOS fica como reserva,
    # para não quebrar quem já tinha preenchido lá
    if "VIDEOS" in wb.sheetnames:
        ws = wb["VIDEOS"]
        porJogo = {g["id"]: g for g in jogos}
        achou = 0
        for r in linhas(ws, acha_cabecalho(ws, "jogo_id")):
            jid = str(r.get("jogo_id") or "").strip()
            vid = str(r.get("video_youtube") or "").strip()
            if not jid or not vid:
                continue
            if jid not in porJogo:
                aviso(f"VIDEOS: jogo '{jid}' não existe na aba JOGOS")
                continue
            # aceita a URL inteira ou só o id
            m = re.search(r"(?:v=|youtu\.be/|embed/)([\w-]{6,})", vid)
            porJogo[jid]["video"] = m.group(1) if m else vid
            porJogo[jid]["videoTitulo"] = str(r.get("titulo") or "").strip()
            achou += 1
        if achou:
            print(f"  {achou} vídeos vindos da aba VIDEOS")

    # ---------------- clipes ----------------
    clipes = []
    if "CLIPES" in wb.sheetnames:
        ws = wb["CLIPES"]
        for r in linhas(ws, acha_cabecalho(ws, "jogador_id")):
            pid = str(r.get("jogador_id") or "").strip()
            video = str(r.get("video_youtube") or "").strip()
            if not pid or not video or pid.upper().startswith("EXEMPLO"):
                continue
            if pid not in porId:
                aviso(f"CLIPES: jogador '{pid}' não existe na aba JOGADORES")
                continue
            clipes.append({
                "jogador": pid,
                "titulo": str(r.get("titulo") or "Melhor momento").strip(),
                "video": video,
                "inicio": num(r.get("inicio_seg")),
                "fim": num(r.get("fim_seg")) or None,
                "temporada": num(r.get("temporada")) or None,
            })

    jogadores.sort(key=lambda p: (-p["jogos"], -p["gols"], p["apelido"]))

    dados = {
        "copa": {"nome": "Copa Revoada", "antigo": "Milior Fut", "playlist": PLAYLIST,
                 "lancesUrl": LANCES_URL},
        "jogadores": jogadores, "times": times, "jogos": jogos,
        "trofeus": trofeus, "escalacoes": escalacoes, "desempenho": desempenho,
        "premiacoes": premiacoes, "clipes": clipes,
    }

    js = "const LOGO='data:image/png;base64,%s';\nconst DADOS=%s;\n" % (
        logo_b64(), json.dumps(dados, ensure_ascii=False, separators=(",", ":")))

    tpl = open(TEMPLATE, encoding="utf-8").read()
    if "// __DADOS__" not in tpl:
        sys.exit("src/app.template.html perdeu o marcador // __DADOS__")
    open(SAIDA, "w", encoding="utf-8").write(tpl.replace("// __DADOS__", js))

    com_foto = sum(1 for p in jogadores if p["foto"])
    com_video = sum(1 for g in jogos if g["video"])
    com_escudo = sum(1 for t in times if t["escudo"])
    com_ft = sum(1 for t in times if t["foto"])
    print(f"\nindex.html gerado — {os.path.getsize(SAIDA)//1024} KB")
    print(f"  {len(jogadores)} jogadores ({com_foto} com foto)")
    print(f"  {len(times)} times ({com_escudo} com escudo, {com_ft} com foto) · "
          f"{len(jogos)} jogos ({com_video} com vídeo)")
    print(f"  {len(escalacoes)} escalações · {len(premiacoes)} troféus · {len(clipes)} clipes")
    print(f"  {len(desempenho)} notas de desempenho por time")

    if not TEM_ROSTO:
        aviso("opencv não está instalado — o recorte do retrato caiu no "
              "enquadramento por proporção. Instale com: pip install "
              "'opencv-python-headless<5'")
    elif ROSTOS_PERDIDOS:
        aviso(f"rosto detectado em {len(ROSTOS_ACHADOS)} fotos; nestas o "
              f"detector não achou e o corte foi por proporção: "
              + ", ".join(ROSTOS_PERDIDOS))
    else:
        print(f"  rosto detectado e centralizado em {len(ROSTOS_ACHADOS)} fotos")

    if not LANCES_URL:
        aviso("LANCES_URL não está definida — a Análise do jogo não grava na "
              "planilha, só na memória da aba. Ver docs/planilha-colaborativa.md")

    ocultos = sorted({e["jogador"] for e in escalacoes if e["oculto"]})
    if ocultos:
        gols_ocultos = sum(e["gols"] for e in escalacoes if e["oculto"])
        print(f"  {len(ocultos)} jogadores escondidos do site seguem contando para o time "
              f"({gols_ocultos} gols): " + ", ".join(ocultos))

    if FUNDO_TIRADO:
        print("  fundo chapado removido do escudo de: " + ", ".join(FUNDO_TIRADO))
    if VAZOU:
        aviso(f"{len(VAZOU)} escudo(s) com desenho claro sobre fundo claro: tirar o "
              "fundo comeria o próprio escudo, então ficaram como vieram. "
              "Se quiser fundo transparente neles, suba o PNG já recortado.")

    sem_num = [p["id"] for p in jogadores if not p["numero"]]
    if sem_num:
        aviso(f"{len(sem_num)} de {len(jogadores)} jogadores sem número de camisa "
              "na coluna numero_camisa da aba JOGADORES. Sem ele, o Estúdio usa a "
              "ordem da lista no lugar do número.")

    sem_foto = [p["id"] for p in jogadores if not p["foto"]]
    if sem_foto:
        aviso(f"{len(sem_foto)} jogadores ainda sem foto: " + ", ".join(sem_foto))
    sobra = assets_sem_uso()
    if sobra:
        aviso("Imagens em assets/ que o build não soube onde encaixar: "
              + ", ".join(sobra) + ". Diga de quem/de que time é cada uma e eu ligo.")

    if AVISOS:
        print("\nAvisos:")
        for a in dict.fromkeys(AVISOS):
            print("  ! " + a)


if __name__ == "__main__":
    main()
